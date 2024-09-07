import os
import sys
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.cell.cell import MergedCell
import itertools




def get_application_path():
    return r'C:\CF'

def get_players_from_excel(num_players):
    app_path = get_application_path()
    excel_file = os.path.join(app_path, "players.xlsx")
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        players = []
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            if row[0] and len(players) < num_players:
                players.append({
                    "name": row[0],
                    "score": 0,
                    "colors": [],
                    "opponent_scores": 0,
                    "wins": 0,
                    "black_wins": 0,
                    "head_to_head": {}
                })
            if len(players) == num_players:
                break
        
        if len(players) < num_players:
            print(f"Warning: Only found {len(players)} players in the Excel file.")
        
        return players
    except FileNotFoundError:
        print(f"Error: 'players.xlsx' file not found in the directory: {app_path}")
        print("Please ensure that the 'players.xlsx' file is in the same directory as the script or executable.")
        input("Press Enter to exit...")
        sys.exit(1)
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {str(e)}")
        print(f"Full path of the file being accessed: {excel_file}")
        input("Press Enter to exit...")
        sys.exit(1)


def create_matches(players, played_matches, round_number):
    if round_number == 1:
        # First round: random pairings
        random.shuffle(players)
    else:
        # Sort players by score, then by opponent's score (for tiebreaks)
        players.sort(key=lambda x: (-x['score'], -x['opponent_scores']))

    matches = []
    paired = set()

    for i, player in enumerate(players):
        if player['name'] in paired:
            continue

        for j in range(i + 1, len(players)):
            opponent = players[j]
            if opponent['name'] in paired:
                continue
            
            if (player['name'], opponent['name']) not in played_matches and \
               (opponent['name'], player['name']) not in played_matches:
                # Determine colors
                if player['colors'].count('W') < opponent['colors'].count('W'):
                    white, black = player, opponent
                elif player['colors'].count('W') > opponent['colors'].count('W'):
                    white, black = opponent, player
                else:
                    # If equal, alternate from last game
                    if player['colors'] and player['colors'][-1] == 'W':
                        white, black = opponent, player
                    else:
                        white, black = player, opponent

                white['colors'].append('W')
                black['colors'].append('B')
                matches.append((white, black))
                paired.add(white['name'])
                paired.add(black['name'])
                break

    # Handle odd number of players
    if len(paired) < len(players):
        unpaired = next(p for p in players if p['name'] not in paired)
        unpaired['colors'].append('X')
        matches.append((unpaired, None))

    return matches

def recommend_rounds(num_players):
    max_rounds = min(int(num_players * 0.4), 12)  # Cap at 12 instead of 10
    if num_players <= 8:
        return min(num_players - 1, max_rounds)
    elif num_players <= 16:
        return min(5, max_rounds)
    elif num_players <= 32:
        return min(6, max_rounds)
    elif num_players <= 64:
        return min(7, max_rounds)
    else:
        return max_rounds

def play_match(matches):
    results = []
    for match in matches:
        if match[1] is None:
            print(f"\n{match[0]['name']} has a bye this round.")
            results.append((match[0], 2, None, 0))
        else:
            print(f"\n{match[0]['name']} vs {match[1]['name']}")
            while True:
                winner = input("Enter the winner (1 for {}, 2 for {}, 'd' for draw, or 'Q' to quit): ".format(match[0]['name'], match[1]['name']))
                if winner.upper() == 'Q':
                    print("Quitting the program...")
                    exit()
                if winner in ['1', '2', 'd']:
                    break
                else:
                    print("Invalid input. Please enter '1', '2', 'd', or 'Q'.")
            
            if winner == '1':
                results.append((match[0], 2, match[1], 0))
            elif winner == '2':
                results.append((match[1], 2, match[0], 0))
            else:
                results.append((match[0], 1, match[1], 1))
    return results

def write_to_excel(players, matches, results, round_number):
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    

    
    # Create 'C:\CF' directory if it doesn't exist
    report_folder = r'C:\CF'
    os.makedirs(report_folder, exist_ok=True)
    
    filename = os.path.join(report_folder, f"game_{current_time}_round_{round_number}.xlsx")

    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet

    colors = [
        'FFC7CE', 'FFEB9C', 'C6EFCE', 'B4C6E7', 'D9D9D9', 'FFA07A', '98FB98', 'ADD8E6',
        'F08080', 'DDA0DD', 'FFE4B5', '20B2AA', 'F0E68C', 'DEB887', '87CEFA', 'D8BFD8',
        'FFDAB9', '7FFFD4', 'F0FFF0', 'FFE4E1', 'E6E6FA', 'FFF0F5', 'F5DEB3', 'FAEBD7',
        'E0FFFF', 'FFB6C1', 'FAFAD2', 'D3D3D3', 'FDF5E6', 'FFEFD5', 'FF69B4', '00FA9A',
        'FF6347', '4682B4', 'FF4500', '9ACD32', 'FF1493', '00CED1', 'FF00FF', '32CD32',
        'FA8072', '4169E1', 'F4A460', '2E8B57', 'F08080', '5F9EA0', 'EE82EE', '6B8E23',
        'FFA500', '483D8B', 'FF7F50', '008080', 'FF6347', '4682B4', 'FF4500', '9ACD32',
        'FF1493', '00CED1', 'FF00FF', '32CD32', 'FA8072', '4169E1'
    ]
    player_colors = {player['name']: PatternFill(start_color=colors[i % len(colors)], end_color=colors[i % len(colors)], fill_type='solid') 
                     for i, player in enumerate(players)}

    ws_pairings = wb.create_sheet(f"Round {round_number} Pairings")
    ws_pairings.title = f"Round {round_number} Pairings"

    headers = ["Table", "White", "Black"]
    for col, header in enumerate(headers, start=1):
        cell = ws_pairings.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for table_num, match in enumerate(matches, start=1):
        ws_pairings.cell(row=table_num+1, column=1, value=table_num)
        player1_cell = ws_pairings.cell(row=table_num+1, column=2, value=match[0]['name'])
        player1_cell.fill = player_colors[match[0]['name']]
        if match[1] is not None:
            player2_cell = ws_pairings.cell(row=table_num+1, column=3, value=match[1]['name'])
            player2_cell.fill = player_colors[match[1]['name']]
        else:
            ws_pairings.cell(row=table_num+1, column=3, value="BYE")

    ws_results = wb.create_sheet(f"Round {round_number} Results")
    ws_results.title = f"Round {round_number} Results"

    headers = ["Table", "White", "Score", "Black", "Score"]
    for col, header in enumerate(headers, start=1):
        cell = ws_results.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for table_num, (match, result) in enumerate(zip(matches, results), start=1):
        ws_results.cell(row=table_num+1, column=1, value=table_num)
        player1_cell = ws_results.cell(row=table_num+1, column=2, value=match[0]['name'])
        player1_cell.fill = player_colors[match[0]['name']]
        ws_results.cell(row=table_num+1, column=3, value=result[1])
        if match[1] is not None:
            player2_cell = ws_results.cell(row=table_num+1, column=4, value=match[1]['name'])
            player2_cell.fill = player_colors[match[1]['name']]
            ws_results.cell(row=table_num+1, column=5, value=result[3])
        else:
            ws_results.cell(row=table_num+1, column=4, value="BYE")
            ws_results.cell(row=table_num+1, column=5, value="-")

    for ws in [ws_pairings, ws_results]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)
    return filename

def update_standings(players, filename, round_number):
    wb = load_workbook(filename)
    sheet_name = f"Standings Round {round_number}"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    headers = ["Rank", "Player", "Score", "Opponent Scores", "Wins", "Black Wins", "Byes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    colors = [
        'FFC7CE', 'FFEB9C', 'C6EFCE', 'B4C6E7', 'D9D9D9', 'FFA07A', '98FB98', 'ADD8E6',
        'F08080', 'DDA0DD', 'FFE4B5', '20B2AA', 'F0E68C', 'DEB887', '87CEFA', 'D8BFD8',
        'FFDAB9', '7FFFD4', 'F0FFF0', 'FFE4E1', 'E6E6FA', 'FFF0F5', 'F5DEB3', 'FAEBD7',
        'E0FFFF', 'FFB6C1', 'FAFAD2', 'D3D3D3', 'FDF5E6', 'FFEFD5', 'FF69B4', '00FA9A',
        'FF6347', '4682B4', 'FF4500', '9ACD32', 'FF1493', '00CED1', 'FF00FF', '32CD32',
        'FA8072', '4169E1', 'F4A460', '2E8B57', 'F08080', '5F9EA0', 'EE82EE', '6B8E23',
        'FFA500', '483D8B', 'FF7F50', '008080', 'FF6347', '4682B4', 'FF4500', '9ACD32',
        'FF1493', '00CED1', 'FF00FF', '32CD32', 'FA8072', '4169E1'
    ]
    player_colors = {player['name']: PatternFill(start_color=colors[i % len(colors)], end_color=colors[i % len(colors)], fill_type='solid') 
                     for i, player in enumerate(players)}

    def custom_sort(player):
        return (player['score'], player['opponent_scores'], player['wins'], player['black_wins'], -player['colors'].count('X'))

    sorted_players = sorted(players, key=custom_sort, reverse=True)
    
    # Apply head-to-head tiebreaker
    for i in range(len(sorted_players) - 1):
        for j in range(i + 1, len(sorted_players)):
            if custom_sort(sorted_players[i]) == custom_sort(sorted_players[j]):
                if head_to_head_compare(sorted_players[i], sorted_players[j]) < 0:
                    sorted_players[i], sorted_players[j] = sorted_players[j], sorted_players[i]

    for rank, player in enumerate(sorted_players, start=1):
        ws.cell(row=rank+1, column=1, value=rank)
        player_cell = ws.cell(row=rank+1, column=2, value=player['name'])
        player_cell.fill = player_colors[player['name']]
        ws.cell(row=rank+1, column=3, value=player['score'])
        ws.cell(row=rank+1, column=4, value=f"{player['opponent_scores']:.2f}")
        ws.cell(row=rank+1, column=5, value=player['wins'])
        ws.cell(row=rank+1, column=6, value=player['black_wins'])
        ws.cell(row=rank+1, column=7, value=player['colors'].count('X'))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)

def get_number_of_players():
    while True:
        try:
            num_players = int(input("Enter the number of players (4-60): "))
            if 4 <= num_players <= 60:
                return num_players
            else:
                print("Please enter a number between 4 and 60.")
        except ValueError:
            print("Please enter a valid number.")

def suggest_rounds(num_players):
    max_rounds = min(num_players - 1, 10)
    suggested_rounds = max_rounds
    if num_players > 20:
        suggested_rounds = min(int(num_players / 2), 10)
    return suggested_rounds, max_rounds

def get_number_of_rounds(suggested_rounds, max_rounds):
    while True:
        try:
            num_rounds = int(input(f"Enter the number of rounds (suggested: {suggested_rounds}, max: {max_rounds}): "))
            if 1 <= num_rounds <= max_rounds:
                return num_rounds
            else:
                print(f"Please enter a number between 1 and {max_rounds}.")
        except ValueError:
            print("Please enter a valid number.")

def display_matches(matches):
    print("\nMatches for this round:")
    for i, match in enumerate(matches, 1):
        if match[1] is None:
            print(f"Table {i}: {match[0]['name']} (White) has a bye")
        else:
            print(f"Table {i}: {match[0]['name']} (White) vs {match[1]['name']} (Black)")

def display_scores(players):
    def custom_sort(player):
        return (player['score'], player['opponent_scores'], player['wins'], player['black_wins'], -player['colors'].count('X'))

    sorted_players = sorted(players, key=custom_sort, reverse=True)
    
    # Apply head-to-head tiebreaker
    for i in range(len(sorted_players) - 1):
        for j in range(i + 1, len(sorted_players)):
            if custom_sort(sorted_players[i]) == custom_sort(sorted_players[j]):
                if head_to_head_compare(sorted_players[i], sorted_players[j]) < 0:
                    sorted_players[i], sorted_players[j] = sorted_players[j], sorted_players[i]

    print("\nCurrent Standings:")
    for i, player in enumerate(sorted_players, 1):
        print(f"{i}. {player['name']}: {player['score']} points (Opponent scores: {player['opponent_scores']:.2f}, Wins: {player['wins']}, Black wins: {player['black_wins']}, Byes: {player['colors'].count('X')})")

def update_scores(results, players):
    for result in results:
        player1, score1, player2, score2 = result
        player1['score'] += score1
        if player2 is not None:
            if score1 == 2:  # A win
                player1['wins'] += 1
                if player1['colors'][-1] == 'B':  # If the player won with black
                    player1['black_wins'] += 1
            player2['score'] += score2
            if score2 == 2:  # A win for the second player
                player2['wins'] += 1
                if player2['colors'][-1] == 'B':  # If the second player won with black
                    player2['black_wins'] += 1
            player1['opponent_scores'] += player2['score']
            player2['opponent_scores'] += player1['score']
            
            # Record head-to-head result
            if score1 > score2:
                player1['head_to_head'][player2['name']] = 'W'
                player2['head_to_head'][player1['name']] = 'L'
            elif score1 < score2:
                player1['head_to_head'][player2['name']] = 'L'
                player2['head_to_head'][player1['name']] = 'W'
            else:
                player1['head_to_head'][player2['name']] = 'D'
                player2['head_to_head'][player1['name']] = 'D'
        else:
            # For byes, the player gets 2 points (a win) and 0 opponent score
            player1['wins'] += 1
            # We don't add anything to opponent_scores for a bye

    # Round all opponent scores to two decimal places
    for player in players:
        player['opponent_scores'] = round(player['opponent_scores'], 2)


def head_to_head_compare(player1, player2):
    if player2['name'] in player1['head_to_head']:
        result = player1['head_to_head'][player2['name']]
        if result == 'W':
            return 1
        elif result == 'L':
            return -1
    return 0  # If they haven't played or drew






def get_application_path():
    if getattr(sys, 'frozen', False):
        # If the application is run as a bundle, the PyInstaller bootloader
        # extends the sys module by a flag frozen=True and sets the app 
        # path into variable _MEIPASS'.
        return sys._MEIPASS
    else:
        return os.path.dirname(os.path.abspath(__file__))

def generate_summary(players, all_matches, all_results, num_rounds, filename):
    # Create 'C:\CF' directory if it doesn't exist
    report_folder = r'C:\CF'
    os.makedirs(report_folder, exist_ok=True)

    # Generate a new filename in the report folder
    base_filename = os.path.basename(filename)
    new_filename = os.path.join(report_folder, f"summary_{base_filename}")

    # Load the workbook and create a new sheet
    wb = load_workbook(filename)
    ws = wb.create_sheet("Tournament Summary")

    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = "Chess Tournament Summary"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ["Rank", "Player"] + [f"Round {i+1}" for i in range(num_rounds)] + ["Total Score", "Opponent Scores", "Wins", "Black Wins", "Byes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Sort players
    def custom_sort(player):
        return (player['score'], player['opponent_scores'], player['wins'], player['black_wins'], -player['colors'].count('X'))

    sorted_players = sorted(players, key=custom_sort, reverse=True)
    
    # Apply head-to-head tiebreaker
    for i in range(len(sorted_players) - 1):
        for j in range(i + 1, len(sorted_players)):
            if custom_sort(sorted_players[i]) == custom_sort(sorted_players[j]):
                if head_to_head_compare(sorted_players[i], sorted_players[j]) < 0:
                    sorted_players[i], sorted_players[j] = sorted_players[j], sorted_players[i]

    # Colors for players
    colors = [
        'FFC7CE', 'FFEB9C', 'C6EFCE', 'B4C6E7', 'D9D9D9', 'FFA07A', '98FB98', 'ADD8E6',
        'F08080', 'DDA0DD', 'FFE4B5', '20B2AA', 'F0E68C', 'DEB887', '87CEFA', 'D8BFD8',
        'FFDAB9', '7FFFD4', 'F0FFF0', 'FFE4E1', 'E6E6FA', 'FFF0F5', 'F5DEB3', 'FAEBD7',
        'E0FFFF', 'FFB6C1', 'FAFAD2', 'D3D3D3', 'FDF5E6', 'FFEFD5', 'FF69B4', '00FA9A',
        'FF6347', '4682B4', 'FF4500', '9ACD32', 'FF1493', '00CED1', 'FF00FF', '32CD32',
        'FA8072', '4169E1', 'F4A460', '2E8B57', 'F08080', '5F9EA0', 'EE82EE', '6B8E23',
        'FFA500', '483D8B', 'FF7F50', '008080', 'FF6347', '4682B4', 'FF4500', '9ACD32',
        'FF1493', '00CED1', 'FF00FF', '32CD32', 'FA8072', '4169E1'
    ]
    player_colors = {player['name']: PatternFill(start_color=colors[i % len(colors)], end_color=colors[i % len(colors)], fill_type='solid') 
                     for i, player in enumerate(players)}

    # Player data
    for rank, player in enumerate(sorted_players, start=1):
        row = rank + 2
        ws.cell(row=row, column=1, value=rank)
        player_cell = ws.cell(row=row, column=2, value=player['name'])
        player_cell.fill = player_colors[player['name']]
        
        # Pairing history and results
        for round in range(num_rounds):
            match = next((m for m in all_matches[round] if player['name'] in [m[0]['name'], m[1]['name'] if m[1] else None]), None)
            result = next((r for r in all_results[round] if player['name'] in [r[0]['name'], r[2]['name'] if r[2] else None]), None)
            
            if match:
                if match[0]['name'] == player['name']:
                    opponent = match[1]['name'] if match[1] else "BYE"
                    color = "W"
                else:
                    opponent = match[0]['name']
                    color = "B"
                
                if result:
                    if result[0]['name'] == player['name']:
                        score = result[1]
                    else:
                        score = result[3]
                    
                    cell_value = f"{color} vs {opponent} ({score})"
                else:
                    cell_value = f"{color} vs {opponent}"
                
                ws.cell(row=row, column=round+3, value=cell_value)
        
        # Total score and other statistics
        ws.cell(row=row, column=num_rounds+3, value=player['score'])
        ws.cell(row=row, column=num_rounds+4, value=f"{player['opponent_scores']:.2f}")
        ws.cell(row=row, column=num_rounds+5, value=player['wins'])
        ws.cell(row=row, column=num_rounds+6, value=player['black_wins'])
        ws.cell(row=row, column=num_rounds+7, value=player['colors'].count('X'))

    # Adjust column widths
    ws.column_dimensions['A'].width = 5  # Set a fixed width for the "Rank" column
    for col in range(2, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(players) + 3):  # +3 to include title and header rows
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue  # Skip merged cells
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=len(players)+2, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Save the workbook with the new filename
    wb.save(new_filename)
    return new_filename

def main():
    report_folder = r'C:\CF'
    print(f"All reports will be saved to: {report_folder}")
    
    num_players = get_number_of_players()
    players = get_players_from_excel(num_players)
    
    recommended_rounds = recommend_rounds(num_players)
    max_rounds = min(int(num_players * 0.4), 12)  # Enforce the 40% rule, capped at 12
    
    print(f"Recommended number of rounds: {recommended_rounds}")
    print(f"Maximum number of rounds: {max_rounds}")
    
    while True:
        try:
            num_rounds = int(input(f"Enter the number of rounds (recommended: {recommended_rounds}, max: {max_rounds}): "))
            if 1 <= num_rounds <= max_rounds:
                break
            else:
                print(f"Please enter a number between 1 and {max_rounds}.")
        except ValueError:
            print("Please enter a valid number.")
    
    played_matches = set()
    filename = None
    all_matches = []
    all_results = []

    for round in range(1, num_rounds + 1):
        print(f"\nRound {round}")
        matches = create_matches(players, played_matches, round)
        all_matches.append(matches)
        
        if not matches:
            print("Unable to create matches. The tournament will end early.")
            break
        
        display_matches(matches)
        
        filename = write_to_excel(players, matches, [], round)
        print(f"Round {round} pairings have been written to {filename}")
        
        input("Press Enter when you're ready to enter the results for this round...")
        
        results = play_match(matches)
        all_results.append(results)
        update_scores(results, players)
        
        filename = write_to_excel(players, matches, results, round)
        print(f"Round {round} results have been written to {filename}")
        
        update_standings(players, filename, round)
        print(f"Standings have been updated in {filename}")
        
        display_scores(players)

        # Update played_matches after each round
        for match in matches:
            if match[1] is not None:
                played_matches.add((match[0]['name'], match[1]['name']))

    print("\nTournament completed. Final Standings:")
    display_scores(players)
    
    # Generate summary
    summary_filename = generate_summary(players, all_matches, all_results, num_rounds, filename)
    print(f"\nTournament summary has been added to {summary_filename}")
    
    print("Thank you for using the Chess Tournament Manager!")

if __name__ == "__main__":
    main()